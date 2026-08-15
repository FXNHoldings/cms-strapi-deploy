#!/usr/bin/env python3
"""
Build a fill-in Excel workbook for supplying phone specifications, and the
converter target for `import-gsmarena-specs-from-file.mjs`.

    python3 scripts/make-specs-workbook.py
    python3 scripts/make-specs-workbook.py --category=tablets --out=/opt/assets/tablets.xlsx

Three sheets:

  Instructions  what each column means and how the pieces fit together
  Products      every product in the category, with the join key already filled
  Specs         the sheet to fill in — one row per specification

`Specs` is deliberately long-form (one row per spec) rather than one column per
spec. GSMArena's sections vary by device — a foldable has rows a candybar does
not — so a fixed column set would either truncate the data or leave most cells
empty. Long form also maps directly onto the importer's nested shape:

    section  ->  specs_json.sections[].section
    label    ->  specs_json.sections[].rows[].label
    value    ->  specs_json.sections[].rows[].value

The key column is `clean_model`, not `product_slug`. 149 products collapse to 66
distinct phone models because specs do not vary by storage capacity, so specs are
entered once per model and the converter fans them out to every product sharing
it.
"""
import argparse
import json
import os
import sys
import urllib.parse
import urllib.request

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

HEADER_FILL = PatternFill("solid", fgColor="1F3864")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
NOTE_FONT = Font(color="7F7F7F", italic=True, size=10)
TITLE_FONT = Font(bold=True, size=13)


def fetch_products(category: str):
    base = os.environ.get("STRAPI_INTERNAL_URL") or os.environ.get("STRAPI_URL") or "http://127.0.0.1:8888"
    token = os.environ.get("STRAPI_API_TOKEN") or os.environ.get("STRAPI_TOKEN") or ""
    params = [
        ("pagination[pageSize]", "1000"),
        ("status", "published"),
        ("filters[tags][$containsi]", "nxt-bargains"),
        ("filters[categories][slug][$eq]", category),
        ("fields[0]", "name"),
        ("fields[1]", "slug"),
    ]
    url = f"{base.rstrip('/')}/api/commerce-products?{urllib.parse.urlencode(params)}"
    req = urllib.request.Request(url, headers={"Authorization": f"Bearer {token}"} if token else {})
    with urllib.request.urlopen(req, timeout=60) as resp:
        return json.load(resp).get("data", [])


def clean_model(name: str) -> str:
    """Strip retail packaging noise a specs database will not carry."""
    import re
    s = name
    s = re.sub(r"\b\d+\s?(gb|tb)\b", "", s, flags=re.I)
    s = re.sub(r"\b(wi-?fi(\s*\+\s*cellular)?|cellular|lte|bluetooth|gps(\s*\+\s*cellular)?)\b", "", s, flags=re.I)
    s = re.sub(r"\b(unlocked|factory unlocked|smartphone|smart phone|prepaid|dual sim|sim free)\b", "", s, flags=re.I)
    s = re.sub(r"\s*\(\s*\)\s*", " ", s)
    s = re.sub(r"\s{2,}", " ", s)
    return s.strip(" ,-")


def style_header(ws, headers, widths):
    for i, (head, width) in enumerate(zip(headers, widths), start=1):
        cell = ws.cell(row=1, column=i, value=head)
        cell.fill, cell.font = HEADER_FILL, HEADER_FONT
        cell.alignment = Alignment(vertical="center")
        ws.column_dimensions[get_column_letter(i)].width = width
    ws.freeze_panes = "A2"


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--category", default="smart-phones")
    ap.add_argument("--out", default=None)
    args = ap.parse_args()

    products = fetch_products(args.category)
    if not products:
        sys.exit(f"No published products in category '{args.category}'.")

    rows = sorted(
        ({"slug": p["slug"], "title": p["name"], "model": clean_model(p["name"])} for p in products),
        key=lambda r: r["title"],
    )
    models = sorted({r["model"] for r in rows})

    out = args.out or f"/opt/assets/{args.category.replace('-', '_')}_specs.xlsx"
    wb = Workbook()

    # ── Instructions ────────────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Instructions"
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 96
    lines = [
        ("How to use this workbook", None),
        ("", None),
        ("Fill in the 'Specs' sheet.", "One row per specification. Leave 'Products' alone — it is reference only."),
        ("", None),
        ("clean_model", "The phone model, e.g. 'Apple iPhone 16 Pro'. This is the key that links your specs "
                        "to products. Copy the values from the 'Products' sheet exactly."),
        ("section", "The spec group, e.g. Display, Platform, Camera, Battery, Body. Becomes one table on the page."),
        ("label", "The spec name within that group, e.g. 'Size', 'Chipset', 'Capacity'."),
        ("value", "The spec value, e.g. '6.9 inches', 'Snapdragon 8 Elite', '5000 mAh'."),
        ("source_url", "Optional. The page the spec came from. Recorded for provenance; the same URL can repeat."),
        ("", None),
        ("Why model, not product?", f"{len(rows)} products in this category share only {len(models)} distinct models, "
                                     "because specifications do not change with storage capacity. Enter each model once "
                                     "and every matching product receives the specs."),
        ("", None),
        ("Order matters", "Rows are grouped into sections in the order they appear here, so keep rows of the same "
                          "section together."),
        ("Blank rows", "Ignored. A row missing clean_model, section, label or value is skipped by the converter."),
        ("", None),
        ("When finished", "Save the file, then ask to convert it. It becomes the JSON that "
                          "import-gsmarena-specs-from-file.mjs loads into Strapi."),
    ]
    r = 1
    for head, body in lines:
        if head:
            c = ws.cell(row=r, column=1, value=head)
            c.font = TITLE_FONT if body is None else Font(bold=True, size=11)
        if body:
            c = ws.cell(row=r, column=2, value=body)
            c.alignment = Alignment(wrap_text=True, vertical="top")
            ws.row_dimensions[r].height = 30
        r += 1

    # ── Products (reference) ────────────────────────────────────────────────
    ws = wb.create_sheet("Products")
    style_header(ws, ["product_slug", "product_title", "clean_model"], [46, 46, 34])
    for i, row in enumerate(rows, start=2):
        ws.cell(row=i, column=1, value=row["slug"])
        ws.cell(row=i, column=2, value=row["title"])
        ws.cell(row=i, column=3, value=row["model"])

    # ── Specs (fill in) ─────────────────────────────────────────────────────
    ws = wb.create_sheet("Specs")
    style_header(ws, ["clean_model", "section", "label", "value", "source_url"], [34, 18, 26, 60, 44])

    # A worked example, so the shape is obvious without reading the instructions.
    example_model = rows[0]["model"]
    example = [
        (example_model, "Body", "Dimensions", "147.6 x 71.6 x 7.8 mm"),
        (example_model, "Body", "Weight", "170 g"),
        (example_model, "Display", "Type", "Super Retina XDR OLED, 120Hz"),
        (example_model, "Display", "Size", "6.1 inches"),
        (example_model, "Platform", "Chipset", "Apple A18"),
        (example_model, "Memory", "Internal", "128GB 8GB RAM"),
        (example_model, "Battery", "Type", "Li-Ion 3561 mAh"),
    ]
    for i, (model, section, label, value) in enumerate(example, start=2):
        ws.cell(row=i, column=1, value=model)
        ws.cell(row=i, column=2, value=section)
        ws.cell(row=i, column=3, value=label)
        ws.cell(row=i, column=4, value=value)
    note = ws.cell(row=len(example) + 2, column=1,
                   value="↑ example rows — overwrite or delete them, then add your own below")
    note.font = NOTE_FONT

    # Dropdown of valid models, so a typo cannot silently orphan a block of specs.
    # Excel caps an inline list at 255 chars, so the values live on a hidden sheet.
    ref = wb.create_sheet("_models")
    for i, m in enumerate(models, start=1):
        ref.cell(row=i, column=1, value=m)
    ref.sheet_state = "hidden"
    dv = DataValidation(
        type="list",
        formula1=f"=_models!$A$1:$A${len(models)}",
        allow_blank=True,
        showDropDown=False,
    )
    dv.error = "Pick a model from the list, or copy one from the Products sheet."
    dv.errorTitle = "Unknown model"
    ws.add_data_validation(dv)
    dv.add(f"A2:A2000")

    wb.save(out)
    print(f"category : {args.category}")
    print(f"products : {len(rows)}")
    print(f"models   : {len(models)} distinct (specs entered once per model)")
    print(f"written  : {out}")


if __name__ == "__main__":
    main()
