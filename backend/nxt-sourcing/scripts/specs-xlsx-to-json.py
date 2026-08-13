#!/usr/bin/env python3
"""
Convert a filled specs workbook into the JSON `import-gsmarena-specs-from-file.mjs`
consumes.

    python3 scripts/specs-xlsx-to-json.py --in /root/smart_phones_specs.xlsx
    python3 scripts/specs-xlsx-to-json.py --in <file> --out data/canonical-products/smartphone_specs.json

Reads the `Specs` sheet (clean_model, section, label, value, source_url) and the
`Products` sheet (product_slug, product_title, clean_model), then emits one
record per *product*.

Specs are entered once per model and fanned out to every product sharing it,
because specifications do not change with storage capacity — 149 phones in this
catalogue are only 66 distinct models. The fan-out happens here rather than in
the spreadsheet so the person filling it in types each spec once.

Section order is preserved as encountered, and rows keep their order within a
section, so the tables render in the order they were entered rather than
alphabetically.

Nothing is invented: a product whose model has no spec rows is omitted from the
output entirely, so the importer never writes an empty spec block over good data.
"""
import argparse
import json
import os
import re
import sys
from collections import OrderedDict, defaultdict
from datetime import date

try:
    from openpyxl import load_workbook
except ImportError:
    sys.exit("openpyxl is required:  pip3 install openpyxl")


def norm(v):
    if v is None:
        return ""
    return re.sub(r"\s+", " ", str(v)).strip()


def read_sheet(wb, name):
    if name not in wb.sheetnames:
        sys.exit(f"Workbook has no '{name}' sheet. Found: {', '.join(wb.sheetnames)}")
    ws = wb[name]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [norm(h).lower() for h in rows[0]]
    out = []
    for raw in rows[1:]:
        record = {headers[i]: norm(v) for i, v in enumerate(raw) if i < len(headers) and headers[i]}
        if any(record.values()):
            out.append(record)
    return out


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--in", dest="src", required=True)
    ap.add_argument("--out", dest="out", default=None)
    ap.add_argument("--source-name", default="GSMArena")
    args = ap.parse_args()

    if not os.path.isfile(args.src):
        sys.exit(f"Input not found: {args.src}")

    wb = load_workbook(args.src, read_only=True, data_only=True)
    spec_rows = read_sheet(wb, "Specs")
    products = read_sheet(wb, "Products")

    if not products:
        sys.exit("The 'Products' sheet is empty — regenerate the workbook.")

    # model -> ordered sections -> ordered rows
    by_model = defaultdict(OrderedDict)
    source_by_model = {}
    skipped = 0
    example_marker = 0

    for row in spec_rows:
        model = row.get("clean_model", "")
        section = row.get("section", "")
        label = row.get("label", "")
        value = row.get("value", "")
        # The note under the example block starts with an arrow, not a model.
        if model.startswith("↑"):
            example_marker += 1
            continue
        if not (model and section and label and value):
            skipped += 1
            continue
        by_model[model].setdefault(section, [])
        by_model[model][section].append({"label": label, "value": value})
        if row.get("source_url"):
            source_by_model.setdefault(model, row["source_url"])

    if not by_model:
        sys.exit("No usable rows on the 'Specs' sheet. Each row needs clean_model, section, label and value.")

    today = date.today().isoformat()
    records = []
    models_used = set()
    products_without_specs = 0

    for p in products:
        slug = p.get("product_slug", "")
        title = p.get("product_title", "")
        model = p.get("clean_model", "")
        if not slug or model not in by_model:
            products_without_specs += 1
            continue
        models_used.add(model)
        sections = [
            {"section": section, "rows": rows}
            for section, rows in by_model[model].items()
        ]
        records.append({
            "product_slug": slug,
            "product_title": title,
            "clean_model": model,
            "gsmarena_url": source_by_model.get(model, ""),
            "status": "matched",
            "match_confidence": "manual",
            "specs_json": {
                "source": args.source_name,
                "source_url": source_by_model.get(model, ""),
                "source_title": model,
                "clean_model": model,
                "extracted_at": today,
                "sections": sections,
            },
        })

    out = args.out or os.path.join(
        os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
        "data/canonical-products/smartphone_specs.json",
    )
    os.makedirs(os.path.dirname(out), exist_ok=True)
    with open(out, "w", encoding="utf-8") as fh:
        json.dump(records, fh, indent=2, ensure_ascii=False)
        fh.write("\n")

    unmatched = sorted(set(by_model) - models_used)

    print(f"spec rows read        : {len(spec_rows)}")
    if skipped:
        print(f"  incomplete, skipped : {skipped}")
    print(f"models with specs     : {len(by_model)}")
    print(f"products written      : {len(records)}")
    print(f"products with no specs: {products_without_specs}")
    if unmatched:
        print(f"\n! {len(unmatched)} model(s) on the Specs sheet match no product — check spelling:")
        for m in unmatched[:10]:
            print(f"    {m}")
    print(f"\nwritten: {out}")
    print("\nNext:")
    print(f"  node scripts/import-gsmarena-specs-from-file.mjs --file={out} --limit=3")
    print(f"  node scripts/import-gsmarena-specs-from-file.mjs --file={out} --write")


if __name__ == "__main__":
    main()
